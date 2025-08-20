import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


def export_to_excel(kpis: dict, region_df: pd.DataFrame, product_df: pd.DataFrame, output_path: str = "reports/sales_report.xlsx"):
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "KPI Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4C72B0")

    ws_summary["A1"] = "KPI"
    ws_summary["B1"] = "Value"
    for cell in [ws_summary["A1"], ws_summary["B1"]]:
        cell.font = header_font
        cell.fill = header_fill

    for i, (k, v) in enumerate(kpis.items(), start=2):
        ws_summary[f"A{i}"] = k.replace("_", " ").title()
        ws_summary[f"B{i}"] = v

    ws_region = wb.create_sheet("Regional Summary")
    for r in dataframe_to_rows(region_df, index=False, header=True):
        ws_region.append(r)

    ws_products = wb.create_sheet("Top Products")
    for r in dataframe_to_rows(product_df, index=False, header=True):
        ws_products.append(r)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"Report saved: {output_path}")
