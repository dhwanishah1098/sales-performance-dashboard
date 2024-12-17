from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

HEADER_FILL  = PatternFill("solid", fgColor="1a1a2e")
HEADER_FONT  = Font(color="FFFFFF", bold=True)
ALT_FILL     = PatternFill("solid", fgColor="f0f4ff")
CENTER       = Alignment(horizontal="center", vertical="center")
THIN         = Side(border_style="thin", color="d0d0d0")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header_row(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
